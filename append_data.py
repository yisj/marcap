import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import shutil
from datetime import datetime


if not os.path.exists("temp"):
	os.mkdir("temp")

download_folder = "C:\\update_xls"
target_file = "data/marcap-2021.csv.gz"
temp_file  = "temp/marcap-2021.csv.gz"

shutil.copy(target_file, temp_file)

column_order = [
	"Code",
	"Name",
	"Market",
	"Dept",
	"Close",
	"ChangeCode",
	"Changes",
	"ChangesRatio",
	"Open",
	"High",
	"Low",
	"Volume",
	"Amount",
	"Marcap",
	"Stocks",
	"MarketId",
	"Rank",
	"Date"
]

def market_id(row):
	if row['Market'] == "KOSPI":
		val = "STK"
	elif row["Market"] == "KOSDAQ":
		val = "KSQ"
	elif row["Market"] == "KONEX":
		val = "KNX"
	else:
		val = ""
	return val

def change_code(row):
	if row["Changes"] > 0:
		val = 1
	elif row["Changes"] < 0:
		val = 2
	else:
		val = 3
	return val

def check_folder(last_date):
	files = list()
	for f in os.listdir(download_folder):
		files.append(f)
	print(files)

	last_date = datetime.strptime(last_date,"%Y-%m-%d")
	files_to_append = list()
	for f in files:
		date = os.path.splitext(f)[0]
		d = datetime.strptime(date,"%Y-%m-%d")
		if (d > last_date):
			files_to_append.append(f)

	print(files_to_append)

	for f in files_to_append:
		date = os.path.splitext(f)[0]
		target_xls = download_folder + "\\" + f
		# df = pd.read_excel(target_xls, sheet_name="Sheet1", keep_default_na=False)
		# print(df)
		# print(df.keys())


		wb = load_workbook(target_xls)
		ws = wb.active

		if ws["E2"].value == "-":
			print(date+" skipped")
			continue

		ws["A1"] = "Code"
		ws["B1"] = "Name"
		ws["C1"] = "Market"
		ws["D1"] = "Dept"
		ws["E1"] = "Close"
		ws["F1"] = "Changes"
		ws["G1"] = "ChangesRatio"
		ws["H1"] = "Open"
		ws["I1"] = "High"
		ws["J1"] = "Low"
		ws["K1"] = "Volume"
		ws["L1"] = "Amount"
		ws["M1"] = "Marcap"
		ws["N1"] = "Stocks"

		data = ws.values
		cols = next(data)[0:]
		data = list(data)[0:]
		print(cols)
		
		df =  pd.DataFrame(data, columns=cols)
		df = df.sort_values(by=["Marcap"], ascending=False)
		df["ChangeCode"] = df.apply(change_code, axis=1)
		df["MarketId"] = df.apply(market_id, axis=1)
		df["Rank"] = np.arange(len(df))
		df["Rank"] += 1
		df["Date"] = date
		df = df[column_order]
		print(df)

		cf_df = pd.read_csv(temp_file)
		if 'ChagesRatio' in cf_df:
			cf_df = cf_df.rename(columns={"ChagesRatio": "ChangesRatio"})
		print(cf_df)
		cf_df = cf_df.append(df)
		cf_df.reset_index(drop=True, inplace=True)
		print(cf_df)


		cf_df.to_csv(temp_file, compression="gzip", index=False)

	# os.remove(temp_file)


def check_last_date():
	df = pd.read_csv(temp_file)
	last_date = df['Date'].values[-1]
	print("Last Date: "+last_date)
	return last_date



if __name__ == "__main__":
	last_date = check_last_date()
	check_folder(last_date)